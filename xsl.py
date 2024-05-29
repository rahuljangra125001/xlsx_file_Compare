import pandas as pd 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Read the Excel files into DataFrames
df1 = pd.read_excel('file1.xlsx')
df2 = pd.read_excel('file2.xlsx')

# Specify the column to compare
column_to_compare = 'Column1'

# Create copies of the original DataFrames to keep track of original indices
df1_orig = df1.copy()
df2_orig = df2.copy()

# Find values only in df1
only_in_df1 = df1[~df1[column_to_compare].isin(df2[column_to_compare])]

# Find values only in df2
only_in_df2 = df2[~df2[column_to_compare].isin(df1[column_to_compare])]

# Load workbooks and select the first sheet
wb1 = load_workbook('file1.xlsx')
ws1 = wb1.active

wb2 = load_workbook('file2.xlsx')
ws2 = wb2.active

# Define fill colors
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

# Function to apply fill color to cells
def apply_fill(ws, df, fill_color, column_to_compare):
    for index, row in df.iterrows():
        # Find the cell to color
        cell = ws.cell(row=index+2, column=df.columns.get_loc(column_to_compare)+1)  # +2 for header and 1-indexed
        cell.fill = fill_color

# Apply red fill to cells in file1 that are not in file2
apply_fill(ws1, only_in_df1, red_fill, column_to_compare)

# Apply red fill to cells in file2 that are not in file1
apply_fill(ws2, only_in_df2, red_fill, column_to_compare)

# Apply yellow fill to cells in file1 that are in both files
apply_fill(ws1, df1[df1[column_to_compare].isin(df2[column_to_compare])], yellow_fill, column_to_compare)

# Apply yellow fill to cells in file2 that are in both files
apply_fill(ws2, df2[df2[column_to_compare].isin(df1[column_to_compare])], yellow_fill, column_to_compare)

# Save the workbooks
wb1.save('file1_comparison.xlsx')
wb2.save('file2_comparison.xlsx')
